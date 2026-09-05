import os
import sys
import time
import json
import base64
import traceback
import smtplib
import shutil
import re
from datetime import datetime, timezone, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import xml.etree.ElementTree as ET
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

# ==========================================
# 📧 Email Notification Credentials & Config
# ==========================================
EMAIL_SENDER = os.environ.get("ALERT_EMAIL_SENDER", "")
EMAIL_PASSWORD = os.environ.get("ALERT_EMAIL_PASSWORD", "")
EMAIL_RECEIVER = os.environ.get("ALERT_EMAIL_RECEIVER", "")

# ==========================================
# 🌐 Google Apps Script Config
# ==========================================
GAS_WEBAPP_URL = os.environ.get("GAS_WEBAPP_URL", "")
GAS_AUTH_KEY = os.environ.get("GAS_AUTH_KEY", "")

EXPORTS_DIR = os.path.abspath("exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

TARGET_URLS = {
    "CMRT": "https://b5.caspio.com/dp/0c4a30006f6c908f547e41cfa9bc",
    "EMRT": "https://c0eku224.caspio.com/dp/0c4a3000f851a3fe32a54dbcbd38",
    "AMRT": "https://c0eku224.caspio.com/dp/0c4a300001be9d377b74464d8a65",
    "REVISIONS": "https://b5.caspio.com/dp/0c4a3000a9ae96d4b36e406fa326",
    "PUBLIC": "https://www.responsiblemineralsinitiative.org/facilities-lists/public-list/",
    "ELIGIBLE": "https://c0eku224.caspio.com/dp/0c4a30001fb4dc1742cd4c88bda8"
}

BASE_TITLE = "RMI Smelter Data Sync"
UUID_PATTERN = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')


def sanitize_traceback(tb_str: str) -> str:
    sanitized = re.sub(r'([A-Za-z]:\\[^:\n\r]+|\/[a-zA-Z0-9_\.\-]+(?:\/[a-zA-Z0-9_\.\-]+)+)', '[INTERNAL_FILE_PATH]',
                       tb_str)
    sanitized = re.sub(r'(auth|password|key|token|secret)[\'"]?\s*[:=]\s*[\'"][^\'"]+[\'"]', r'\1: "***MASKED***"',
                       sanitized, flags=re.IGNORECASE)
    return sanitized


def send_daily_email_report(subject: str, body_text: str):
    if not all([EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECEIVER]):
        print("\n⚠️ [Email Notification Skipped]: Missing email credentials.")
        return

    try:
        msg = MIMEMultipart()
        msg["From"] = f"RMI Smelter Sync Bot <{EMAIL_SENDER}>"
        msg["To"] = EMAIL_RECEIVER
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        print(f"\n📧 [Email Report Sent Successfully] Receiver: {EMAIL_RECEIVER}")
    except Exception as ex:
        print(f"\n❌ [Email Delivery Failed]: {ex}")


def cleanup_local_temp_files():
    if not os.path.exists(EXPORTS_DIR):
        return
    cleaned = 0
    for filename in os.listdir(EXPORTS_DIR):
        file_path = os.path.join(EXPORTS_DIR, filename)
        if os.path.isfile(file_path) and UUID_PATTERN.match(filename):
            try:
                os.remove(file_path)
                cleaned += 1
            except Exception:
                pass
    if cleaned > 0:
        print(f"🧹 [Auto-Cleanup] Cleaned {cleaned} temporary Playwright download file(s).")


def purge_all_local_exports():
    if not os.path.exists(EXPORTS_DIR):
        return
    deleted_count = 0
    for filename in os.listdir(EXPORTS_DIR):
        file_path = os.path.join(EXPORTS_DIR, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.remove(file_path)
                deleted_count += 1
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
                deleted_count += 1
        except Exception:
            pass
    if deleted_count > 0:
        print(f"🔒 [Security Complete] Cleaned {deleted_count} file(s) from local exports.")


def download_caspio_direct(page, target_name, url):
    save_path = os.path.join(EXPORTS_DIR, f"{target_name}.xml")
    print(f"[{target_name}] Requesting live XML export from Caspio DataPage...")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        time.sleep(2)

        btn = page.locator(
            "a.cbResultSetDownloadLink, a[data-cb-name='DataDownloadButton'], a:has-text('Download Data')").first
        btn.wait_for(state="attached", timeout=25000)

        try:
            btn.scroll_into_view_if_needed(timeout=3000)
        except Exception:
            pass

        time.sleep(1)

        with page.expect_download(timeout=45000) as download_info:
            btn.click(force=True)
            time.sleep(1)

            opt = page.locator("a:has-text('Excel(XML)'), div:has-text('Excel(XML)'), li:has-text('Excel(XML)')").last
            if opt.is_visible(timeout=5000):
                opt.click(force=True)
            else:
                try:
                    opt.wait_for(state="attached", timeout=3000)
                    opt.click(force=True)
                except Exception:
                    page.keyboard.press("Enter")

        download = download_info.value
        download.save_as(save_path)
        size_kb = os.path.getsize(save_path) / 1024
        print(f"   -> ✅ [{target_name}] Downloaded: {size_kb:.1f} KB")
    except Exception as e:
        print(f"   -> ❌ [{target_name}] Failed: {e}")
        raise e


def handle_rmi_public_export(page, url):
    print(f"\n[PUBLIC] Navigating to portal: {url}")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        time.sleep(2)

        try:
            cookie_btn = page.locator(
                "button.btn-close, .cookie-close, [aria-label='Close'], button:has-text('✕')").first
            if cookie_btn.is_visible(timeout=2000):
                cookie_btn.click(force=True)
                time.sleep(1)
        except Exception:
            pass

        try:
            accept_btn = page.locator(
                "input[value='I Accept'], input[value*='Accept'], button:has-text('Accept')").first
            if accept_btn.is_visible(timeout=3000):
                accept_btn.click(force=True)
                print("   -> [PUBLIC] Terms accepted ('I Accept' clicked)")
                time.sleep(2)
        except Exception:
            pass

        page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        print("[PUBLIC] Searching for 'Download Excel' button...")
        excel_btn = None
        for _ in range(30):
            for frame in page.frames:
                candidate = frame.locator(
                    "input[value='Download Excel'], button:has-text('Download Excel'), a:has-text('Download Excel')").first
                try:
                    if candidate.is_visible(timeout=1000):
                        excel_btn = candidate
                        break
                except Exception:
                    continue
            if excel_btn:
                break
            time.sleep(1)

        if not excel_btn:
            raise Exception("Could not locate 'Download Excel' button on the public list page.")

        with page.expect_download(timeout=60000) as download_info:
            try:
                excel_btn.evaluate("el => el.click()")
            except Exception:
                excel_btn.click(force=True)

        download = download_info.value
        suggested_name = download.suggested_filename
        ext = os.path.splitext(suggested_name)[1].lower() or ".xlsx"
        save_path = os.path.join(EXPORTS_DIR, f"PUBLIC{ext}")
        download.save_as(save_path)

        size_kb = os.path.getsize(save_path) / 1024
        print(f"   -> ✅ [PUBLIC] Downloaded as '{os.path.basename(save_path)}' ({size_kb:.1f} KB)")
        return save_path

    except Exception as e:
        print(f"   -> ❌ [PUBLIC] Failed: {e}")
        raise e


def run_live_pipeline():
    print("=========================================================")
    print(" 🚀 Phase 1: Automated Live Data Harvesting")
    print("=========================================================")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, downloads_path=EXPORTS_DIR)
        context = browser.new_context(
            accept_downloads=True,
            ignore_https_errors=True,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        context.add_cookies([
            {"name": "rmiViewAgree", "value": "true", "domain": ".responsiblemineralsinitiative.org", "path": "/"},
            {"name": "cb_disclaimer_agreed", "value": "true", "domain": ".caspio.com", "path": "/"}
        ])

        page = context.new_page()

        for name in ["CMRT", "EMRT", "AMRT", "REVISIONS", "ELIGIBLE"]:
            download_caspio_direct(page, name, TARGET_URLS[name])
            time.sleep(1)

        handle_rmi_public_export(page, TARGET_URLS["PUBLIC"])
        time.sleep(2)

        browser.close()

    cleanup_local_temp_files()


def parse_spreadsheet_ml(xml_path):
    if not os.path.exists(xml_path) or os.path.getsize(xml_path) < 100:
        return []
    tree = ET.parse(xml_path)
    root = tree.getroot()
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    grid = []
    for row in root.findall(".//ss:Row", ns):
        row_cells = []
        col_idx = 0
        for cell in row.findall("./ss:Cell", ns):
            idx_attr = cell.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
            if idx_attr:
                target_idx = int(idx_attr) - 1
                while col_idx < target_idx:
                    row_cells.append("")
                    col_idx += 1
            data_elem = cell.find("./ss:Data", ns)
            val = data_elem.text if data_elem is not None and data_elem.text else ""
            row_cells.append(val.strip())
            col_idx += 1
        if any(row_cells):
            grid.append(row_cells)
    return grid


def parse_flexible_grid(filepath):
    if not os.path.exists(filepath):
        return []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        grid = []
        for row in sheet.iter_rows(values_only=True):
            row_vals = [str(c).strip() if c is not None else "" for c in row]
            if any(row_vals):
                grid.append(row_vals)
        wb.close()
        if grid:
            return grid
    except Exception:
        pass

    try:
        grid = parse_spreadsheet_ml(filepath)
        if grid:
            return grid
    except Exception:
        pass

    return []


def find_col_idx(headers, keywords):
    for kw in keywords:
        clean_kw = "".join(filter(str.isalnum, kw)).lower()
        for i, h in enumerate(headers):
            if not h:
                continue
            clean_h = "".join(filter(str.isalnum, str(h))).lower()
            if clean_kw in clean_h:
                return i
    return -1


def format_date(val):
    val_str = str(val).strip()
    if len(val_str) >= 10 and val_str[:4].isdigit() and val_str[4] == "-" and val_str[7] == "-":
        return val_str[:10]
    return val_str


def send_gas_request_with_retry(payload: dict, context_name: str, max_retries: int = 3, initial_delay: int = 6) -> dict:
    last_error_text = ""
    last_status = 0

    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(
                GAS_WEBAPP_URL,
                headers={"Content-Type": "text/plain;charset=utf-8"},
                data=json.dumps(payload),
                timeout=60,
                allow_redirects=True
            )
            last_status = resp.status_code
            last_error_text = resp.text

            if resp.status_code in [404, 429, 500, 502, 503, 504]:
                if attempt < max_retries:
                    wait_sec = initial_delay * attempt
                    print(
                        f"   ⚠️ [{context_name}] Status {resp.status_code} (Google Transient Error). Retrying in {wait_sec}s ({attempt}/{max_retries})...")
                    time.sleep(wait_sec)
                    continue
                else:
                    raise Exception(
                        f"[{context_name}] Failed after {max_retries} attempts. Status: {last_status}, Response: {last_error_text[:300]}")

            resp_json = {}
            try:
                resp_json = resp.json()
            except Exception:
                pass

            if resp.status_code == 200 and resp_json.get("status") == "success":
                return resp_json
            else:
                if attempt < max_retries:
                    wait_sec = initial_delay * attempt
                    print(
                        f"   ⚠️ [{context_name}] Non-success response: {resp.text[:120]}. Retrying in {wait_sec}s ({attempt}/{max_retries})...")
                    time.sleep(wait_sec)
                    continue
                else:
                    raise Exception(
                        f"[{context_name}] GAS returned error. Status: {resp.status_code}, Response: {resp.text}")

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                wait_sec = initial_delay * attempt
                print(
                    f"   ⚠️ [{context_name}] Network/Timeout Exception: {e}. Retrying in {wait_sec}s ({attempt}/{max_retries})...")
                time.sleep(wait_sec)
            else:
                raise e

    raise Exception(f"[{context_name}] All {max_retries} attempts exhausted.")


def log_summary_to_gas_history(today_str, original_source_counts):
    """구글 스프레드시트의 'Summary History' 탭에 당일 원본 카운트를 1행 누적 기록합니다."""
    if not GAS_WEBAPP_URL:
        print("⚠️ GAS_WEBAPP_URL is missing. Cannot record summary history.")
        return

    payload = {
        "action": "record_summary_history",
        "auth": GAS_AUTH_KEY,
        "record": {
            "date": today_str,
            "cmrt": original_source_counts["CMRT"],
            "emrt": original_source_counts["EMRT"],
            "amrt": original_source_counts["AMRT"],
            "revision": original_source_counts["Revision"],
            "eligible": original_source_counts["Eligible"],
            "public": original_source_counts["Public"]
        }
    }

    try:
        send_gas_request_with_retry(payload, context_name="Record Summary History", max_retries=3, initial_delay=4)
        print(f"   -> 📈 [Summary History Logged]: Successfully added row to Google Sheet 'Summary History'.")
    except Exception as e:
        print(f"   ⚠️ Could not record summary history to GAS: {e}")


def consolidate_and_export(output_filename, timestamp_full_str, today_str):
    print("\n=========================================================")
    print(" 📊 Phase 2: Data Parsing, RMAP Mapping & Consolidation")
    print("=========================================================")

    base_rows = []
    public_facility_map = {}
    eligible_facility_map = {}
    revisions_map = {}

    # 원본 파일 내 고유 제련소/시설 레코드 수 카운트
    original_source_counts = {
        "CMRT": 0,
        "EMRT": 0,
        "AMRT": 0,
        "Revision": 0,
        "Eligible": 0,
        "Public": 0
    }

    # 1. RMI Public List 파싱
    pub_candidates = [
        os.path.join(EXPORTS_DIR, f) for f in os.listdir(EXPORTS_DIR)
        if f.startswith("PUBLIC.") or f.startswith("PUBLIC_LIST.")
    ]
    if not pub_candidates:
        raise ValueError("PUBLIC download file not found in exports directory.")

    public_file_path = pub_candidates[0]
    public_grid = parse_flexible_grid(public_file_path)
    if not public_grid:
        raise ValueError(f"Failed to parse public facilities list: {public_file_path}")

    header_row_idx = 0
    for idx, row in enumerate(public_grid[:5]):
        if find_col_idx(row, ["facilityid", "smelterid"]) != -1:
            header_row_idx = idx
            break

    pub_headers = public_grid[header_row_idx]
    p_metal_idx = find_col_idx(pub_headers, ["metal"])
    p_id_idx = find_col_idx(pub_headers, ["facilityid", "cid", "smelterid"])
    p_name_idx = find_col_idx(pub_headers, ["standardfacilityname", "standardsmeltername", "facilityname"])
    p_op_status_idx = find_col_idx(pub_headers, ["facilityoperationalstatus", "operationalstatus"])
    p_level_idx = find_col_idx(pub_headers, ["supplychainlevel"])
    p_country_idx = find_col_idx(pub_headers, ["countrylocation", "country"])
    p_rmap_status_idx = find_col_idx(pub_headers,
                                     ["assessmentprogramstatus", "duediligenceassessmentprogramstatus", "programstatus",
                                      "rmapstatus"])
    p_cycle_idx = find_col_idx(pub_headers, ["duediligenceassessmentcycle", "assessmentcycle"])
    p_audit_date_idx = find_col_idx(pub_headers, ["lastonsiteassessmentdate", "lastaudit", "auditdate"])
    p_reaudit_idx = find_col_idx(pub_headers, ["reassessmentinprogress", "reaudit"])

    for r in public_grid[header_row_idx + 1:]:
        cid = r[p_id_idx].strip() if p_id_idx != -1 and p_id_idx < len(r) and r[p_id_idx] else ""
        if not cid:
            continue

        public_facility_map[cid] = {
            "metal": r[p_metal_idx].strip() if p_metal_idx != -1 and p_metal_idx < len(r) and r[p_metal_idx] else "",
            "name": r[p_name_idx].strip() if p_name_idx != -1 and p_name_idx < len(r) and r[p_name_idx] else "",
            "op_status": r[p_op_status_idx].strip() if p_op_status_idx != -1 and p_op_status_idx < len(r) and r[
                p_op_status_idx] else "",
            "level": r[p_level_idx].strip() if p_level_idx != -1 and p_level_idx < len(r) and r[p_level_idx] else "",
            "country": r[p_country_idx].strip() if p_country_idx != -1 and p_country_idx < len(r) and r[
                p_country_idx] else "",
            "rmap_status": (r[p_rmap_status_idx].strip() if p_rmap_status_idx != -1 and p_rmap_status_idx < len(r) and
                                                            r[p_rmap_status_idx] else "") or "-",
            "cycle": r[p_cycle_idx].strip() if p_cycle_idx != -1 and p_cycle_idx < len(r) and r[p_cycle_idx] else "",
            "audit_date": format_date(r[p_audit_date_idx]) if p_audit_date_idx != -1 and p_audit_date_idx < len(
                r) else "",
            "reaudit": (r[p_reaudit_idx].strip() if p_reaudit_idx != -1 and p_reaudit_idx < len(r) and r[
                p_reaudit_idx] else "") or "No"
        }
    original_source_counts["Public"] = len(public_facility_map)
    print(f"• Facilities in original RMI Public List: {original_source_counts['Public']} records")

    # 2. RMI Eligible Facilities List 파싱
    elg_candidates = [
        os.path.join(EXPORTS_DIR, f) for f in os.listdir(EXPORTS_DIR)
        if f.startswith("ELIGIBLE.") or f.startswith("ELIGIBLE_LIST.")
    ]
    if elg_candidates:
        elg_file_path = elg_candidates[0]
        elg_grid = parse_flexible_grid(elg_file_path)
        if elg_grid:
            elg_hdr_idx = 0
            for idx, row in enumerate(elg_grid[:5]):
                if find_col_idx(row, ["facilityid", "cid", "smelterid"]) != -1:
                    elg_hdr_idx = idx
                    break
            elg_headers = elg_grid[elg_hdr_idx]
            e_metal_idx = find_col_idx(elg_headers, ["metal"])
            e_id_idx = find_col_idx(elg_headers, ["facilityid", "cid", "smelterid"])
            e_name_idx = find_col_idx(elg_headers, ["standardfacilityname", "facilityname"])
            e_level_idx = find_col_idx(elg_headers, ["supplychainlevel", "level"])
            e_country_idx = find_col_idx(elg_headers, ["countrylocation", "country"])
            e_state_idx = find_col_idx(elg_headers, ["stateprovinceregion", "state", "province"])

            for r in elg_grid[elg_hdr_idx + 1:]:
                cid = r[e_id_idx].strip() if e_id_idx != -1 and e_id_idx < len(r) and r[e_id_idx] else ""
                if not cid:
                    continue
                eligible_facility_map[cid] = {
                    "metal": r[e_metal_idx].strip() if e_metal_idx != -1 and e_metal_idx < len(r) and r[
                        e_metal_idx] else "",
                    "name": r[e_name_idx].strip() if e_name_idx != -1 and e_name_idx < len(r) and r[e_name_idx] else "",
                    "level": r[e_level_idx].strip() if e_level_idx != -1 and e_level_idx < len(r) and r[
                        e_level_idx] else "Pinch Point",
                    "country": r[e_country_idx].strip() if e_country_idx != -1 and e_country_idx < len(r) and r[
                        e_country_idx] else "",
                    "state": r[e_state_idx].strip() if e_state_idx != -1 and e_state_idx < len(r) and r[
                        e_state_idx] else ""
                }
            original_source_counts["Eligible"] = len(eligible_facility_map)
            print(f"• Facilities in original RMI Eligible List: {original_source_counts['Eligible']} records")

    # 3. Revision History 파싱
    rev_grid = parse_spreadsheet_ml(os.path.join(EXPORTS_DIR, "REVISIONS.xml"))
    if not rev_grid:
        raise ValueError("REVISIONS.xml parsing failed.")
    headers_rev = rev_grid[0]
    metal_idx_rev = find_col_idx(headers_rev, ["metal"])
    id_idx_rev = find_col_idx(headers_rev, ["smelterid", "cid", "facilityid"])
    name_idx_rev = find_col_idx(headers_rev,
                                ["standardsmeltername", "standardfacilityname", "smeltername", "facilityname"])
    country_idx_rev = find_col_idx(headers_rev, ["country"])
    basis_idx_rev = find_col_idx(headers_rev, ["basisforrevision", "basis", "revision"])
    details_idx_rev = find_col_idx(headers_rev, ["details", "comments", "history"])
    date_idx_rev = find_col_idx(headers_rev, ["revisiondate", "revdate", "date"])

    for r in rev_grid[1:]:
        s_id = r[id_idx_rev].strip() if id_idx_rev != -1 and id_idx_rev < len(r) and r[id_idx_rev] else ""
        if s_id:
            metal = r[metal_idx_rev].strip() if metal_idx_rev != -1 and metal_idx_rev < len(r) and r[
                metal_idx_rev] else ""
            name = r[name_idx_rev].strip() if name_idx_rev != -1 and name_idx_rev < len(r) and r[name_idx_rev] else ""
            country = r[country_idx_rev].strip() if country_idx_rev != -1 and country_idx_rev < len(r) and r[
                country_idx_rev] else ""
            basis = r[basis_idx_rev].strip() if basis_idx_rev != -1 and basis_idx_rev < len(r) and r[
                basis_idx_rev] else ""
            details = r[details_idx_rev].strip() if details_idx_rev != -1 and details_idx_rev < len(r) and r[
                details_idx_rev] else ""
            rev_date = format_date(r[date_idx_rev]) if date_idx_rev != -1 and date_idx_rev < len(r) else ""
            info = f"{basis}: {details}" if basis and details else (basis or details or "-")

            if s_id not in revisions_map or rev_date >= revisions_map[s_id]["date"]:
                revisions_map[s_id] = {
                    "metal": metal,
                    "name": name,
                    "country": country,
                    "info": info,
                    "date": rev_date
                }
    original_source_counts["Revision"] = len(revisions_map)
    print(f"• Unique facilities in Revision History: {original_source_counts['Revision']} records")

    # 4. CMRT / EMRT / AMRT 템플릿 데이터 로드
    for t_name in ["CMRT", "EMRT", "AMRT"]:
        t_grid = parse_spreadsheet_ml(os.path.join(EXPORTS_DIR, f"{t_name}.xml"))
        if not t_grid:
            raise ValueError(f"{t_name}.xml parsing failed.")
        headers_t = t_grid[0]
        metal_idx = find_col_idx(headers_t, ["metal"])
        ref_idx = find_col_idx(headers_t, ["smelterreference", "reference"])
        name_idx = find_col_idx(headers_t, ["standardsmeltername", "standardfacilityname", "smeltername"])
        country_idx = find_col_idx(headers_t, ["country"])
        id_idx = find_col_idx(headers_t, ["smelterid", "cid", "facilityid"])
        city_idx = find_col_idx(headers_t, ["city"])
        state_idx = find_col_idx(headers_t, ["stateprovince", "state", "province"])

        count_in_type = 0
        for r in t_grid[1:]:
            cid_val = r[id_idx].strip() if id_idx != -1 and id_idx < len(r) and r[id_idx] else ""
            if not cid_val and not (r[name_idx].strip() if name_idx != -1 and name_idx < len(r) else ""):
                continue
            base_rows.append({
                "type": t_name,
                "metal": r[metal_idx].strip() if metal_idx != -1 and metal_idx < len(r) and r[metal_idx] else "",
                "smelterRef": r[ref_idx].strip() if ref_idx != -1 and ref_idx < len(r) and r[ref_idx] else "",
                "facilityName": r[name_idx].strip() if name_idx != -1 and name_idx < len(r) and r[name_idx] else "",
                "country": r[country_idx].strip() if country_idx != -1 and country_idx < len(r) and r[
                    country_idx] else "",
                "cid": cid_val,
                "city": r[city_idx].strip() if city_idx != -1 and city_idx < len(r) and r[city_idx] else "",
                "state": r[state_idx].strip() if state_idx != -1 and state_idx < len(r) and r[state_idx] else "",
            })
            count_in_type += 1
        original_source_counts[t_name] = count_in_type
        print(f"• Facilities in original {t_name} Reference List: {count_in_type} records")

    headers_out = [
        "No.", "Source", "Metal", "CID", "Operation Status", "Level", "CAHRA",
        "Standard Facility Name", "Country", "Smelter Reference", "City",
        "State Province", "RMAP Status", "Audit / Cycle / Reaudit", "Revision History"
    ]

    all_table_data = []
    processed_ids = set()
    conformant_matched_count = 0
    active_matched_count = 0
    row_counter = 1

    # 5-1. 베이스 템플릿(CMRT/EMRT/AMRT) 머지
    for item in base_rows:
        cid = item["cid"]
        country = item["country"]
        op_status = ""
        level = "Pinch Point"
        rmap_status = "-"
        audit_info = ""

        if cid and cid in eligible_facility_map:
            level = eligible_facility_map[cid]["level"] or level

        if cid and cid in public_facility_map:
            pub_info = public_facility_map[cid]
            op_status = pub_info["op_status"]
            level = pub_info["level"] or level
            if not country:
                country = pub_info["country"]
            rmap_status = pub_info["rmap_status"]

            if "conform" in rmap_status.lower():
                conformant_matched_count += 1
                audit_info = f"{pub_info['audit_date']} / {pub_info['cycle']} / {pub_info['reaudit']}"
            elif "active" in rmap_status.lower() or "participat" in rmap_status.lower():
                active_matched_count += 1

        rev_history = revisions_map[cid]["info"] if cid and cid in revisions_map else ""

        all_table_data.append([
            row_counter,
            item["type"],
            item["metal"],
            cid,
            op_status,
            level,
            "",
            item["facilityName"],
            country,
            item["smelterRef"],
            item["city"],
            item["state"],
            rmap_status,
            audit_info,
            rev_history
        ])
        if cid:
            processed_ids.add(cid)
        row_counter += 1

    # 5-2. Eligible Facilities List 추가 머지
    for elg_cid, elg_val in eligible_facility_map.items():
        if elg_cid not in processed_ids:
            country = elg_val["country"]
            op_status = ""
            level = elg_val["level"] or "Upstream"
            rmap_status = "-"
            audit_info = ""

            if elg_cid in public_facility_map:
                pub_info = public_facility_map[elg_cid]
                op_status = pub_info["op_status"]
                level = pub_info["level"] or level
                if not country:
                    country = pub_info["country"]
                rmap_status = pub_info["rmap_status"]

                if "conform" in rmap_status.lower():
                    conformant_matched_count += 1
                    audit_info = f"{pub_info['audit_date']} / {pub_info['cycle']} / {pub_info['reaudit']}"
                elif "active" in rmap_status.lower() or "participat" in rmap_status.lower():
                    active_matched_count += 1

            rev_history = revisions_map[elg_cid]["info"] if elg_cid in revisions_map else ""

            all_table_data.append([
                row_counter,
                "Eligible",
                elg_val["metal"],
                elg_cid,
                op_status,
                level,
                "",
                elg_val["name"],
                country,
                "",
                "",
                elg_val["state"],
                rmap_status,
                audit_info,
                rev_history
            ])
            processed_ids.add(elg_cid)
            row_counter += 1

    # 5-3. Public List 단독 시설 추가 머지
    for pub_cid, pub_val in public_facility_map.items():
        if pub_cid not in processed_ids:
            country = pub_val["country"]
            rmap_status = pub_val["rmap_status"]
            audit_info = ""

            if "conform" in rmap_status.lower():
                conformant_matched_count += 1
                audit_info = f"{pub_val['audit_date']} / {pub_val['cycle']} / {pub_val['reaudit']}"
            elif "active" in rmap_status.lower() or "participat" in rmap_status.lower():
                active_matched_count += 1

            rev_history = revisions_map[pub_cid]["info"] if pub_cid in revisions_map else ""

            all_table_data.append([
                row_counter,
                "Public",
                pub_val["metal"],
                pub_cid,
                pub_val["op_status"],
                pub_val["level"],
                "",
                pub_val["name"],
                country,
                "",
                "",
                "",
                rmap_status,
                audit_info,
                rev_history
            ])
            processed_ids.add(pub_cid)
            row_counter += 1

    # 5-4. Revision History (삭제된 제련소) 머지
    removed_count = 0
    for rev_id, rev_val in revisions_map.items():
        if rev_id not in processed_ids:
            country = rev_val["country"]

            all_table_data.append([
                row_counter,
                "Revision",
                rev_val["metal"],
                rev_id,
                "",
                "",
                "",
                rev_val["name"],
                country,
                "",
                "",
                "",
                "Removed",
                "",
                rev_val["info"] or "Removed"
            ])
            processed_ids.add(rev_id)
            removed_count += 1
            row_counter += 1

    total_facilities = len(all_table_data)
    standard_count = total_facilities - conformant_matched_count - active_matched_count - removed_count

    summary_stats = {
        "total": total_facilities,
        "cmrt": original_source_counts["CMRT"],
        "emrt": original_source_counts["EMRT"],
        "amrt": original_source_counts["AMRT"],
        "revision": original_source_counts["Revision"],
        "eligible": original_source_counts["Eligible"],
        "public": original_source_counts["Public"],
        "conformant": conformant_matched_count,
        "active": active_matched_count,
        "standard": standard_count,
        "removed": removed_count,
        "timestamp": timestamp_full_str
    }

    # 6. 마스터 엑셀 워크북 빌드 (당일 단일 요약 1행 구성)
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Disclaimer & Summary"

    sum_headers = ["Data Consolidated", "CMRT", "EMRT", "AMRT", "Revision", "Eligible", "Public"]
    sum_values = [
        today_str,
        original_source_counts["CMRT"],
        original_source_counts["EMRT"],
        original_source_counts["AMRT"],
        original_source_counts["Revision"],
        original_source_counts["Eligible"],
        original_source_counts["Public"]
    ]

    for col_idx, h_text in enumerate(sum_headers, start=2):
        ws_summary.cell(row=2, column=col_idx, value=h_text)
    for col_idx, val in enumerate(sum_values, start=2):
        cell = ws_summary.cell(row=3, column=col_idx, value=val)
        if isinstance(val, (int, float)):
            cell.number_format = "#,##0"

    font_summary_header = Font(name="Pretendard", size=11, bold=True, color="1E293B")
    fill_summary_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_summary_body = Font(name="Pretendard", size=11)
    align_center = Alignment(horizontal="center", vertical="center")

    thin_side = Side(style="thin", color="CBD5E1")
    box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws_summary.row_dimensions[2].height = 24
    ws_summary.row_dimensions[3].height = 22

    for col in range(2, 9):
        c_head = ws_summary.cell(row=2, column=col)
        c_head.font = font_summary_header
        c_head.fill = fill_summary_header
        c_head.alignment = align_center
        c_head.border = box_border

        c_val = ws_summary.cell(row=3, column=col)
        c_val.font = font_summary_body
        c_val.alignment = align_center
        c_val.border = box_border

    # 하단 신규 안내 문구 및 Disclaimer (5행 고정 배치)
    disclaimer_text = (
        "a2MDS Consulting\n"
        "글로벌 제품환경규제 대응 전문기업\n"
        "IMDS | Responsible·Conflict Minerals | Product Environmental Compliance | Supply Chain Due Diligence\n"
        "APA Engineering과의 전략적 파트너십을 기반으로, 교육부터 컨설팅, 아웃소싱, 자동화 솔루션까지 One-stop으로 지원합니다.\n\n"
        "Disclaimer\n"
        "본 자료는 RMI(Responsible Minerals Initiative) 웹사이트에서 제공하는 시설 및 제련소 목록을 기반으로 작성되었습니다.\n"
        "본 자료의 정보는 자료 송부일 이전에 확인된 내용을 기준으로 합니다.\n"
        "RMI 목록은 지속적으로 업데이트되므로, 본 자료의 작성일 이후 변경된 최신 정보와 차이가 있을 수 있습니다.\n"
        "따라서 본 자료는 통합 목록 예시로 활용하여 주시고, 최신 정보가 필요한 경우 RMI 공식 웹사이트에서 최신 제련소 및 시설 정보를 직접 확인하시기 바랍니다.\n\n"
        "RMI 제련소 및 시설 정보\n"
        "• 링크: https://www.responsiblemineralsinitiative.org/\n"
        "• 사용된 목록 정보: Smelter Reference List (CMRT, EMRT, AMRT, Revision), RMI Eligible Facilities List, RMI Public Facilities List"
    )

    ws_summary.merge_cells("B5:H5")
    cell_disclaimer = ws_summary.cell(row=5, column=2, value=disclaimer_text)
    cell_disclaimer.font = Font(name="Pretendard", size=11, color="1E293B")
    cell_disclaimer.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_summary.row_dimensions[5].height = 360

    summary_widths = {1: 4, 2: 24, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14}
    for col_idx, width in summary_widths.items():
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    # Facility Log 탭 빌드
    ws_log = wb.create_sheet(title="Facility Log")
    ws_log.append(headers_out)
    for r_data in all_table_data:
        ws_log.append(r_data)

    font_body = Font(name="Pretendard", size=11)
    font_header = Font(name="Pretendard", size=11, bold=True, color="1E293B")
    fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    align_header = Alignment(horizontal="center", vertical="center")
    align_body = Alignment(vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="medium", color="94A3B8")
    )

    ws_log.row_dimensions[1].height = 26
    for cell in ws_log[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = thin_border

    for row in ws_log.iter_rows(min_row=2):
        for cell in row:
            cell.font = font_body
            cell.alignment = align_body

    ws_log.freeze_panes = "E2"
    ws_log.auto_filter.ref = ws_log.dimensions

    custom_widths = [6, 10, 12, 13, 16, 14, 10, 28, 16, 22, 14, 16, 16, 34, 38]
    for i, w in enumerate(custom_widths, 1):
        ws_log.column_dimensions[get_column_letter(i)].width = w

    output_filepath = os.path.join(EXPORTS_DIR, f"{output_filename}.xlsx")
    wb.save(output_filepath)
    wb.close()

    print(f"\n✨ Master Excel File Generated: {output_filepath}")
    return output_filepath, summary_stats, headers_out, all_table_data, original_source_counts


def upload_file_via_gas(filepath, filename, mime_type):
    if not GAS_WEBAPP_URL:
        print("⚠️ GAS_WEBAPP_URL is missing. Cannot upload file.")
        return

    try:
        with open(filepath, "rb") as f:
            encoded_bytes = base64.b64encode(f.read()).decode("utf-8")

        payload = {
            "action": "upload_file",
            "auth": GAS_AUTH_KEY,
            "fileName": filename,
            "mimeType": mime_type,
            "fileData": encoded_bytes
        }

        send_gas_request_with_retry(payload, context_name=f"Upload {filename}", max_retries=3, initial_delay=6)
        print(f"   -> ⬆️ [GAS Uploaded]: {filename}")
    except Exception as e:
        print(f"   -> ❌ GAS File Upload Exception: {e}")


def sync_to_google_services(excel_filepath, headers, rows_data):
    print("\n=========================================================")
    print(" ☁️ Phase 3: Syncing Files & Live Google Spreadsheet")
    print("=========================================================")

    VALID_EXTENSIONS = ('.xml', '.xlsx')
    current_local_files = [
        f for f in os.listdir(EXPORTS_DIR)
        if os.path.isfile(os.path.join(EXPORTS_DIR, f))
           and f.lower().endswith(VALID_EXTENSIONS)
           and not UUID_PATTERN.match(f)
    ]

    print(f"📦 Total files to sync to Google Drive ({len(current_local_files)} files): {current_local_files}")
    for fname in current_local_files:
        fpath = os.path.join(EXPORTS_DIR, fname)
        if fname.endswith('.xlsx'):
            mtype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mtype = 'application/xml'
        upload_file_via_gas(fpath, fname, mtype)

    if not GAS_WEBAPP_URL:
        raise ValueError("GAS_WEBAPP_URL environment variable is missing. Cannot sync to Google Spreadsheet.")

    print("\n   -> 📊 Updating Google Spreadsheet via Apps Script Live DB...")
    CHUNK_SIZE = 500
    total_rows = len(rows_data)
    total_chunks = (total_rows + CHUNK_SIZE - 1) // CHUNK_SIZE

    kst_now_str = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M:%S")

    try:
        for i in range(total_chunks):
            start = i * CHUNK_SIZE
            end = min(start + CHUNK_SIZE, total_rows)
            chunk = rows_data[start:end]

            payload = {
                "action": "save_smelters_chunk",
                "auth": GAS_AUTH_KEY,
                "isFirstChunk": (i == 0),
                "lastUpdated": kst_now_str,
                "headers": headers if (i == 0) else [],
                "rows": chunk
            }

            send_gas_request_with_retry(
                payload,
                context_name=f"Chunk {i + 1}/{total_chunks}",
                max_retries=3,
                initial_delay=6
            )

            print(f"   -> ⏳ Synced chunk ({i + 1}/{total_chunks}) to Live Sheet...")

        print("   -> ✅ [Live Sheet Updated]: Successfully synced master data and refreshed Latest Harvest time!")
    except Exception as e:
        print(f"\n❌ Live Sheet Update Failed: {e}")
        traceback.print_exc(file=sys.stdout)
        raise e

    print("\n✅ Google Drive & Live Sync Completed Successfully!")


if __name__ == "__main__":
    kst = timezone(timedelta(hours=9))
    now_kst = datetime.now(kst)
    today_str = now_kst.strftime("%Y-%m-%d")
    today_file_tag = now_kst.strftime("%Y%m%d")
    timestamp_full_str = now_kst.strftime("%Y-%m-%d %H:%M:%S") + " KST (UTC+9)"

    base_name = f"{BASE_TITLE}_{today_file_tag}"

    print(f"\n=== RMI Facility & Smelter Daily Sync Started at {timestamp_full_str} ===")

    try:
        run_live_pipeline()
        excel_path, stats, headers, rows_data, raw_counts = consolidate_and_export(base_name, timestamp_full_str,
                                                                                   today_str)

        # 구글 시트 동기화: 1) Smelter Log 탭 업데이트 및 2) Summary History 탭 1행 누적
        sync_to_google_services(excel_path, headers, rows_data)
        log_summary_to_gas_history(today_str, raw_counts)

        success_subject = f"✅ [SUCCESS] RMI Smelter & Facility Daily Sync Report ({today_file_tag})"
        success_body = (
            f"Dear Mr. CEO,\n\n"
            f"The daily automated harvesting, multi-tier supply chain consolidation, and cloud database synchronization have been successfully completed.\n\n"
            f"==================================================\n"
            f" 📌 DAILY SOURCE & CONSOLIDATION SUMMARY ({today_str})\n"
            f"==================================================\n"
            f"1. Date consolidated: {stats['timestamp']}\n"
            f"2. Original Source Counts (Raw File):\n"
            f"   - CMRT (3TG)                  : {stats['cmrt']:,}\n"
            f"   - EMRT (Cobalt/Mica)          : {stats['emrt']:,}\n"
            f"   - AMRT (Aluminum)             : {stats['amrt']:,}\n"
            f"   - Revision History            : {stats['revision']:,}\n"
            f"   - Eligible List               : {stats['eligible']:,}\n"
            f"   - RMI Public List             : {stats['public']:,}\n\n"
            f"3. Consolidated Master Database  : {stats['total']:,} records\n"
            f"   - Conformant                  : {stats['conformant']:,}\n"
            f"   - Active                      : {stats['active']:,}\n"
            f"   - Standard (-)                : {stats['standard']:,}\n"
            f"   - Removed                     : {stats['removed']:,}\n\n"
            f"• Cloud & Database Synchronization\n"
            f"   - Master File                 : {base_name}.xlsx\n"
            f"   - Google Drive Archive        : Updated\n"
            f"   - Live Sheet Database         : Synced & Latest Harvest Timestamp Refreshed\n"
            f"   - Summary History Tab         : Daily Record Appended\n"
            f"==================================================\n"
        )
        send_daily_email_report(success_subject, success_body)

    except Exception as e:
        error_trace = sanitize_traceback(traceback.format_exc())
        print("\n" + "=" * 57)
        print(" ❌ PIPELINE ERROR OCCURRED")
        print("=" * 57)
        print(f"Error Type: {type(e).__name__}")
        print(f"Error Message: {str(e)}\n")
        print("Detailed Traceback (Sanitized):")
        print(error_trace)
        print("=" * 57 + "\n")

        fail_subject = f"🚨 [FAILURE] RMI Smelter & Facility Sync Error Alert ({today_file_tag})"
        fail_body = (
            f"Dear Mr. CEO,\n\n"
            f"An error occurred during the daily automated synchronization pipeline. The operation has been halted.\n\n"
            f"==================================================\n"
            f" ❌ ERROR SUMMARY\n"
            f"==================================================\n"
            f"• Error Type    : {type(e).__name__}\n"
            f"• Error Message : {str(e)}\n\n"
            f"==================================================\n"
            f" 🔍 SANITIZED TRACEBACK\n"
            f"==================================================\n"
            f"==================================================\n"
            f"{error_trace}\n"
            f"==================================================\n"
            f"※ You can forward this entire error traceback directly to ReS for prompt analysis and troubleshooting."
        )
        send_daily_email_report(fail_subject, fail_body)
        sys.exit(1)
     finally:
        purge_all_local_exports()
